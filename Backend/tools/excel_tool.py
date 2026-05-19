import json
import os
import tempfile
import time
import pandas as pd
from langchain_core.tools import tool
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

@tool
def generate_excel(json_string: str) -> str:
    """
    Generates a professional Excel spreadsheet from business research data.
    Input MUST be a valid JSON string.
    Returns the generated file path.
    """

    try:
        clean_str = json_string.strip().replace("```json", "").replace("```", "")
        data = json.loads(clean_str)

        file_name = f"Research_Report_{int(time.time())}.xlsx"
        file_path = os.path.join(tempfile.gettempdir(), file_name)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:

            # =========================
            # Overview Sheet
            # =========================
            overview_data = {
                "Business Idea": [data.get("businessIdea", "")],
                "Unique Selling Proposition": [data.get("UniqueSellingProposition", "")],
                "Confidence": [data.get("confidence", "")]
            }

            overview_df = pd.DataFrame(overview_data)
            overview_df.to_excel(writer, sheet_name="Overview", index=False)

            # =========================
            # Competitors Sheet
            # =========================
            if "competitors" in data and data["competitors"]:
                competitors_df = pd.DataFrame(data["competitors"])
                competitors_df.to_excel(writer, sheet_name="Competitors", index=False)

            # =========================
            # Trends & Keywords Sheet
            # =========================
            trends = data.get("marketTrends", []) or []
            keywords = data.get("topKeywordsforSEO", []) or []

            max_len = max(len(trends), len(keywords), 1)

            trends_padded = trends + [""] * (max_len - len(trends))

            kw_names = [k.get("name", "") for k in keywords]
            kw_names += [""] * (max_len - len(kw_names))

            kw_vol = [k.get("searchVolume", "") for k in keywords]
            kw_vol += [""] * (max_len - len(kw_vol))

            trends_df = pd.DataFrame({
                "Market Trends": trends_padded,
                "SEO Keyword": kw_names,
                "Search Volume": kw_vol
            })

            trends_df.to_excel(writer, sheet_name="Trends & Keywords", index=False)

            # =========================
            # Workbook Styling
            # =========================
            workbook = writer.book

            # Colors
            header_fill = PatternFill(
                start_color="1E293B",
                end_color="1E293B",
                fill_type="solid"
            )

            alternate_fill = PatternFill(
                start_color="F8FAFC",
                end_color="F8FAFC",
                fill_type="solid"
            )

            header_font = Font(
                color="FFFFFF",
                bold=True,
                size=12
            )

            body_font = Font(
                size=11,
                color="111827"
            )

            border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            # =========================
            # Apply Styling
            # =========================
            for sheet in workbook.worksheets:

                # Freeze header row
                sheet.freeze_panes = "A2"

                # Style all cells
                for row_idx, row in enumerate(sheet.iter_rows(), start=1):

                    for cell in row:

                        cell.border = border
                        cell.font = body_font

                        # Wrap text for readability
                        cell.alignment = Alignment(
                            wrap_text=True,
                            vertical="top",
                            horizontal="left"
                        )

                        # Header row
                        if row_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(
                                horizontal="center",
                                vertical="center",
                                wrap_text=True
                            )

                        # Alternate row color
                        elif row_idx % 2 == 0:
                            cell.fill = alternate_fill

                # =========================
                # Auto-adjust column widths
                # =========================
                for column_cells in sheet.columns:

                    max_length = 0
                    column_letter = column_cells[0].column_letter

                    for cell in column_cells:
                        try:
                            cell_value = str(cell.value) if cell.value else ""
                            max_length = max(max_length, len(cell_value))
                        except:
                            pass

                    adjusted_width = min(max_length + 5, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # =========================
                # Adjust row heights dynamically
                # =========================
                for row in sheet.iter_rows():
                    max_lines = 1

                    for cell in row:
                        if cell.value:
                            lines = len(str(cell.value)) // 40 + 1
                            max_lines = max(max_lines, lines)

                    sheet.row_dimensions[row[0].row].height = max_lines * 18

        return file_path

    except Exception as e:
        return f"Excel generation failed: {str(e)}"